# -*- coding: utf-8 -*-
"""
Created on Mon Oct  4 16:53:47 2021

@author: Theo.ROSSI
"""


class Tracker():
    
    def __init__(self):
        
        self.fig = fig
        self.ax1 = ax1
        self.ax2 = ax2
        self.ax3 = ax3
        self.ax4 = ax4
        self.ax5 = ax5
        
        self.index, self.index2 = [],[]
        self.idx_fback = []
        self.f_back_values = []
        self.list_bouton = []
        self.ind = 0
        
        files = sorted(os.listdir(Path))
        
        LIST_FRAME = []
        
        
        for file in files:

            if 'ROI.tif' in file:   
                
                self.ROI_file = plt.imread('{}/{}'.format(Path, file))
                self.ROI_file = self.ROI_file[:,:,0]
                self.im0 = ax1.imshow(self.ROI_file)
                
                continue
                
            elif '.coord' in file:
                
                Coord_file = pd.read_csv("{}/{}".format(Path, file), sep="\t", header=1)
                
                X, Y = [],[]
                for i in range(len(Coord_file)): 
                    
                    basic_value = Coord_file.iloc[i,0]
                    
                    X.append(int(basic_value.split(',')[0]))
                    Y.append(int(basic_value.split(',')[1]))
                
                coordinates = {'X':X, 'Y':Y}
                self.df_coord = pd.DataFrame(coordinates)
                
                continue
                
            elif '.ini' in file:
                
                ini = '{}/{}'.format(Path, file)
        
                config = cp.ConfigParser()
                config.read(ini)
                
                self.sampling_rate, self.px_dwell_time = float(config.get('_', 'lines.per.second')), float(config.get('_', 'pixel.dwell.time.in.sec'))*1000
            
            else:
                
                data = plt.imread('{}/{}'.format(Path, file)).transpose(1,0,2)
                data = data[:,1:,0]
                LIST_FRAME.append(data)
        
        self.time_vector = np.arange(0, 1.0/self.sampling_rate*len(data[1]), 1.0/self.sampling_rate)
        
        self.fback_start = np.ravel(np.where(self.time_vector >= 0))[0]
        self.fback_stop = np.ravel(np.where(self.time_vector <= 0.1))[-1]
        
        self.frame = np.stack(LIST_FRAME).transpose(1,2,0)
        
        self.slices = self.frame.shape[2]
        
        self.im = ax2.imshow(self.frame[:, :, self.ind])
        
        self.update()




    def markers(self, event):
        
        if event.inaxes in [ax1]:
            
            for idx in range(self.df_coord.shape[0]):
                
                if int(event.xdata) == self.df_coord.iloc[idx][0] and int(event.ydata) == self.df_coord.iloc[idx][1]:
                    
                    self.index = idx
                    self.xdata = int(event.xdata)
                    self.ydata = int(event.ydata)
                    
                    self.bouton = np.mean(self.frame[self.index - 5 : self.index + 5, :, self.ind], axis=0)
                    
                    self.ax1.scatter(self.xdata, self.ydata, s=30, marker='o')
                    self.ax2.scatter(15, self.index, s=100, marker='>')
                    
                    self.ax1.annotate('Idx: {}'.format(self.index),
                                      xy=(self.xdata, self.ydata),
                                      xytext=(self.xdata, self.ydata - 5), color='w')
                    self.ax2.annotate('{}'.format(self.index), xy=(15,self.index), xytext=(15, self.index - 8), color='w')
                        
                    self.ax4.plot(self.time_vector, self.bouton, alpha=0.5)
                    
                        
        if event.inaxes in [ax2]:
            
            if event.button == 2:
                
                self.index2.append(int(event.ydata))
                
                self.bouton = np.mean(self.frame[self.index2[-1]-5 : self.index2[-1]+5, :, self.ind], axis=0)
                self.list_bouton.append(self.bouton)
                
                self.ax2.scatter(15, self.index2[-1], s=100, marker='>')
                self.ax4.plot(self.time_vector, self.bouton, alpha=0.5)
                
                
                if len(self.list_bouton) == self.slices:
                    
                    dataframe_traces = pd.DataFrame(self.list_bouton, index=None, columns=None).transpose()
                    dataframe_traces['Time'] = self.time_vector
                    DATAFRAME_TRACES.append(dataframe_traces)
                    
                    self.index2.insert(0, self.index)
                    dataframe_data = pd.concat((pd.DataFrame(self.index2).transpose(), pd.DataFrame(self.idx_fback).transpose(), pd.DataFrame(self.f_back_values).transpose()), axis=0)
                    dataframe_data.index = ['Trace_index', 'Fback_index', 'Fback_value']
                    DATAFRAME_DATA.append(dataframe_data)
                    print(dataframe_traces)
                    print(dataframe_data)

                
            if event.button == 3:
                
                self.idx_fback.append(int(event.ydata))
                
                self.f_back_trace = np.mean(self.frame[self.idx_fback[-1]-2 : self.idx_fback[-1]+2, :, self.ind], axis=0)
                self.f_back_value = np.mean(self.f_back_trace[self.fback_start : self.fback_stop])
                
                if len(self.f_back_values) >= 1:
                    
                    self.f_back_values.append(self.f_back_value)
                
                self.ax2.scatter(15, int(event.ydata), s=100, marker='$f$')
                self.ax3.plot(self.time_vector, self.f_back_trace, alpha=0.2)
                self.ax3.scatter(self.time_vector[-1]/8, self.f_back_value, marker='_', s=200)
    
    
    
    
    def presskey(self, event):
        
        if event.key == 'd':
            
            self.idx_fback.pop()
        
        
        elif event.key == 'c':
            
            self.ax1.clear()
            self.ax2.clear()
            self.ax3.clear()
            self.ax4.clear()
            self.ax5.clear()
            
            self.im0 = ax1.imshow(self.ROI_file)
            
            self.xdata = []
            self.ydata = []
            self.index = []
            self.index2.clear()
            self.idx_fback.clear()
            self.f_back_values.clear()
            
            self.bouton = []
            self.list_bouton.clear()
            
            self.im = ax2.imshow(self.frame[:, :, self.ind])
            
            DATAFRAME_TRACES.clear()
            DATAFRAME_DATA.clear()
            
        
        elif event.key == 'enter':
            
            self.ax3.clear()
            
            self.idx_fback = [self.idx_fback[-1]]
            self.f_back_trace = np.mean(self.frame[self.idx_fback[-1]-2 : self.idx_fback[-1]+2, :, self.ind], axis=0)
            self.f_back_value = np.mean(self.f_back_trace[self.fback_start : self.fback_stop])
            self.f_back_values.append(self.f_back_value)
            
            self.bouton = np.mean(self.frame[self.index - 5 : self.index + 5, :, self.ind], axis=0)
            self.list_bouton.append(self.bouton)
            
            self.ax3.plot(self.time_vector, self.f_back_trace, alpha=0.4, label='Raw')
            self.ax5.plot(self.time_vector, self.bouton, alpha=0.4)
            
        
        elif event.key == 'm':
            
            self.index2.pop()
            self.list_bouton.pop()

        
        elif event.key == 'a':
            
            dataframe_traces = pd.DataFrame(self.list_bouton, index=None, columns=None).transpose()
            dataframe_traces['Time'] = self.time_vector
            DATAFRAME_TRACES.append(dataframe_traces)
            
            self.index2.insert(0, self.index)
            dataframe_data = pd.concat((pd.DataFrame(self.index2).transpose(), pd.DataFrame(self.idx_fback).transpose(), pd.DataFrame(self.f_back_values).transpose()), axis=0)
            dataframe_data.index = ['Trace_index', 'Fback_index', 'Fback_value']
            DATAFRAME_DATA.append(dataframe_data)
            print(dataframe_traces)
            print(dataframe_data)
                    
        
                
            
            
    def onscroll(self, event):
        
        if event.button == 'up':
            
            self.ind = (self.ind + 1) % self.slices
            
        else:
            
            self.ind = (self.ind - 1) % self.slices

        self.update()

        
        
        
    def update(self):
        
        self.ax1.clear()
        self.ax2.clear()
        self.ax3.clear()
        self.ax4.clear()
        self.ax5.clear()
        
        self.im0 = ax1.imshow(self.ROI_file)
        self.im = ax2.imshow(self.frame[:, :, self.ind])
        self.coord = ax1.plot(self.df_coord['X'], self.df_coord['Y'], color='darkred', lw='1')
        
        if self.index != []:
            self.ax1.scatter(self.xdata, self.ydata, s=30, marker='o')
            self.ax1.annotate('Idx: {}'.format(self.index), xy=(self.xdata, self.ydata), xytext=(self.xdata, self.ydata - 5), color='w')
           
            self.ax2.scatter(15, self.index, s=100, marker='>')
            self.ax2.annotate('{}'.format(self.index), xy=(15,self.index), xytext=(15,self.index - 8), color='w')
        
        
        [self.ax2.scatter(15, self.index2[idx2], s=100, marker='>') for idx2 in range(len(self.index2))]
        
        
        if self.idx_fback != []:
                    
           
            self.f_back_trace = np.mean(self.frame[self.idx_fback[-1]-2 : self.idx_fback[-1]+2, :, self.ind], axis=0)
            self.f_back_value = np.mean(self.f_back_trace[self.fback_start : self.fback_stop])
            
            self.ax2.scatter(15, self.idx_fback[-1], s=100, marker='$f$')
            self.ax3.plot(self.time_vector, self.f_back_trace, alpha=0.2)
            
            self.ax3.scatter(self.time_vector[-1]/8, self.f_back_value, marker='_', s=200)
            
            
        if self.index != []:
                
            self.ax4.plot(self.time_vector, self.bouton, alpha=0.5)
            self.ax4.set_title('Bouton index {}'.format(self.index))
            
            
        if self.list_bouton != []:
            
            for profile in range(len(self.list_bouton)):
                    
                self.ax5.plot(self.time_vector, self.list_bouton[profile], alpha=0.3)

        
        self.ax1.set_title('{}'.format(Path.split('/')[-1]), fontsize=8)
        self.ax2.set_title('Linescan #{}/{}'.format(self.ind, self.slices-1), fontsize=10)
        self.ax3.set_title('Background noise')
        self.ax3.grid(axis='y', linestyle='--')
        self.ax4.set_ylabel('A.U.')
        self.ax4.grid(axis='y', linestyle='--')
        self.ax5.set_title('All profiles')
        self.ax5.set_xlabel('Time (sec)')
        self.ax5.set_ylabel('A.U.')
        self.ax5.grid(axis='y', linestyle='--')
        self.ax5.legend('{}'.format(len(self.list_bouton)))
        
        self.im.axes.figure.canvas.draw()
       

           
if __name__ == '__main__':
    
    import matplotlib.pyplot as plt
    import matplotlib.gridspec as gridspec
    import PySimpleGUI as sg
    import configparser as cp
    import numpy as np
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl import load_workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    import os
     
    DATAFRAME_TRACES = []
    DATAFRAME_DATA = []
    
    sg.theme('DarkBlue')
    
    layout = [[sg.Text('Folder Name')],
              [sg.InputText(size=(35,1), key='Path'), sg.FolderBrowse()],
              [sg.Button('Start Folder')],
              [sg.Frame(layout=[
              [sg.Text('Saving path'), sg.InputText(size=(20,1), key='save'), sg.FolderBrowse()],
              [sg.Text('Date'), sg.InputText(size=(9,1), key='date'), sg.Text('Linescan#'), sg.InputText(size=(4,1), key='linescan')],
              [sg.Text('Freq'), sg.InputText(size=(3,1), key='freq'), sg.Text('#Pulses'), sg.InputText(size=(3,1), key='pulses')],
              [sg.Text('[Ca2+]'), sg.InputText(size=(4,1), key='calcium'), sg.Text('Bouton#'), sg.InputText(size=(3,1), key='bouton')]], title='Excel file name', relief=sg.RELIEF_SUNKEN)],
              [sg.Button('To Excel')],
              [sg.Button('Clear', button_color=('white','darkred'))]]
              
    window = sg.Window('PROFILES EXTRACTION', layout, location=(0,0))
    
    while True:
        event, value = window.read()
        
        try:
            if event in (None, 'Close'):
                plt.close()
                DATAFRAME_TRACES.clear()
                break
            
            if event == 'Start Folder':
                
                Path = value['Path']
                
                fig = plt.figure(figsize=(19,10))
                gs = fig.add_gridspec(3,3)
                ax1 = fig.add_subplot(gs[:,0])
                ax2 = fig.add_subplot(gs[:,1])
                ax3 = fig.add_subplot(gs[0,2])
                ax4 = fig.add_subplot(gs[1,2])
                ax5 = fig.add_subplot(gs[2,2])
                
                tracker = Tracker()
                fig.canvas.mpl_connect('scroll_event', tracker.onscroll)
                fig.canvas.mpl_connect('button_release_event', tracker.markers)
                fig.canvas.mpl_connect('key_release_event', tracker.presskey)
            
            
            if event == 'To Excel':
        
                filepath_traces = str(value['date'])+'_linescan'+str(value['linescan'])+'_'+str(value['freq'])+'Hz_'+str(value['pulses'])+'pulses_'+str(value['calcium'])+'mMCa_'+'bouton'+str(value['bouton'])+'_traces'
                
                if DATAFRAME_TRACES != [] and DATAFRAME_DATA != []:
                    
                    wb = Workbook()
    
                    sheet1 = wb.create_sheet('Traces')
                    sheet2 = wb.create_sheet('Data')
                        
                    [sheet1.append(i) for i in dataframe_to_rows(DATAFRAME_TRACES[0], index=False, header=True)]
                    [sheet2.append(i) for i in dataframe_to_rows(DATAFRAME_DATA[0], index=['Trace_index', 'Fback_index', 'Fback_value'], header=True)]
                
                    print('----------------------')
                
                    wb.remove(wb['Sheet'])
                
                    wb.save('{}\{}.xlsx'.format(value['save'], filepath_traces))
                    
                    
                    sg.popup_ok('Profiles have been stored in:\n{}'.format(filepath_traces))
                    
                
                    DATAFRAME_TRACES.clear()
                    DATAFRAME_DATA.clear()
                    
                else:
                    
                    sg.popup_error('DATAFRAME empty')
    
            if event == 'Clear':
                
                Path = []
                DATAFRAME_TRACES.clear()
                DATAFRAME_DATA.clear()
                plt.close()
        
        except Exception as ex:
            template = 'An exception of type {0} occured. Arguments:\n{1!r}'
            message = template.format(type(ex).__name__, ex.args)
            print(message)
    
    window.close()