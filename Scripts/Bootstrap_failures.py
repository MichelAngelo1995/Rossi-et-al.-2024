# -*- coding: utf-8 -*-
"""
Created on Tue Mar 23 10:14:04 2021

@author: Theo.ROSSI
"""


def gauss_func(mu,sigma,bins):
    '''
    mu : mean
    sigma : standard deviation
    bins : histogram bins

    Returns
    -------
    y : gaussian function of a histogram dataset

    '''
    y = ((1 / (np.sqrt(2 * np.pi) * sigma)) * np.exp(-0.5 * (1 / sigma * (bins - mu))**2)) 
    return y



def func_mono_exp(x, a, b, c, d):
    return a * np.exp(-(x-b)/c) + d



def MAD(a,axis=None):
     '''
     Computes median absolute deviation of an array along given axis
     '''
     #Median along given axis but keep reduced axis so that result can still broadcast along a

     med = np.nanmedian(a, axis=axis, keepdims=True)
     mad = np.nanmedian(np.abs(a-med),axis=axis) #MAD along the given axis
     return mad
 
  
    
def bootstrap_patterns(patterns, run=5000, N=12, input_method='average', output_method='average'):
    
    '''
    Bootstraps synaptic patterns and returns median or average pattern
    
    patterns (list of arrays) : the patterns (data)
    run (int) : the amount of runs for average/median
    N (int) : number of draws for each cycle
    input_method (str) : 'median' , 'average' stores median or average value for each run 
    output_method (str) : 'median' or 'average' : returns medianed or averaged pattern
    
    '''
    
    endCycle = []
    
    for i in range(run): 
        
        temp = []
        
        for j in range(N):
        
            randIndex = np.random.randint(0,len(patterns),size=1)[0]
                        
            temp.append(patterns[randIndex])
            
            if len(temp) == N:
                pass
            else:
                continue
                
        if input_method == 'median' : 
            endCycle.append(np.nanmedian(temp, axis=0))
        
        elif input_method == 'average' : 
            endCycle.append(np.nanmean(temp, axis=0))


    if output_method == 'median': 
        out_bootstrap = np.nanmedian(endCycle, axis=0) 
        out_deviation = MAD(endCycle, axis=0)
        
    elif output_method == 'average': 
        out_bootstrap = np.nanmean(endCycle, axis=0)
        out_deviation = np.nanstd(endCycle, axis=0)
        
    return np.asarray(out_bootstrap), np.asarray(out_deviation), endCycle



def load_xls(file_xls):
    '''
    Loads excel file and returns the timescale, the averaged trace and the list of sweeps
    /!\ The excel file must contain :
        A column for each sweep
        A column for the average
        A column for the time
    '''
    
    df = pd.read_excel (file_xls, sheet_name='Traces DF_F0', header = 0) #sheet_name = 'Traces DF_F0'
    filt = 9
    
    SWEEPS = []
    for i in range(len(df.columns)):
        if 'Time' == df.columns[i]:
            timescale = df.iloc[:,i].values
        
        elif 'Average' == df.columns[i]:
            average = savgol_filter(df.iloc[:,i].values, filt, 2)
            # average = df.iloc[:,i].values
        
        else:
            sweep = savgol_filter(df.iloc[:,i].values, filt, 2)
            # sweep = df.iloc[:,i].values
            SWEEPS.append(sweep)
            
    return timescale, average, SWEEPS



def Fit_single_trace(time, trace, x_start, x_end):
    '''
    Computes an exponential fit on a window of a dataset values.
    The dataset has to be an excel file with variables as columns.
    
    time : time variable.
    trace : the trace the fit must be applied on.
    x_start : first limit of the window.
    x_end : second limit of the window.

    Returns
    -------
    popt : array
        optimal values for the parameters
    idx_start : int
        the first index of the window.
    idx_stop : int
        the second index of the window.

    '''
    
    idx_start = np.ravel(np.where(time >= x_start))[0]
    idx_stop = np.ravel(np.where(time <= x_end))[-1]
    
    x = time[idx_start:idx_stop]
    y = trace[idx_start:idx_stop]
    x2 = np.array(np.squeeze(x))
    y2 = np.array(np.squeeze(y))  
    
    try:
        param_bounds=([-np.inf,0.,0.,-1000.],[np.inf,1.,10.,1000.])      # be careful ok for seconds. If millisec change param 2 and 3
        popt, pcov = curve_fit(func_mono_exp, x2, y2, bounds=param_bounds, maxfev=10000) 
        return popt[2], popt, idx_start, idx_stop
    except:
        print ('Fit failed')
        popt[2]= float('nan')
        popt= float('nan')
        return popt[2], popt, idx_start, idx_stop
        pass



def bleaching_correction(time, trace, start, stop):
    '''
    Suppresses the exponential decay on a window of a dataset values
    
    time (array): time variable.
    trace (array or list): the trace or list of traces the fit must be applied on.
    start (int or float) : first limit of the window.
    end (int or float): second limit of the window.

    Returns
    -------
    no_bleach : list
        The list of traces corrected for the exp decay.

    '''
    
    x1 = float(start)
    x2 = float(stop)
    
    if len(trace) == 0:
        print('Bleaching failed: empty window')
        
    elif type(trace) == np.ndarray:
        tau, popt, xstart, xstop = Fit_single_trace(time, trace, x1, x2)
        bleach = func_mono_exp(time, *popt)
        no_bleach = trace-bleach

    elif type(trace) == list:
        param = [Fit_single_trace(time, trace[i], x1, x2) for i in range(len(trace))]
        bleach = [func_mono_exp(time, *param[i][1]) for i in range(len(trace))]
        no_bleach = [trace[i]-bleach[i] for i in range(len(trace))]
    return no_bleach
        
    
        
def leak_subtraction(time, trace, start, stop):
    ''' 
    Suppresses the offset on a window of a dataset values
    time (array): time variable.
    trace (array or list): the trace or list of traces.
    start (int or float) : first limit of the window.
    end (int or float): second limit of the window.

    Returns
    -------
    no_leak : list
        the list of traces without offset.

    '''
    
    x1 = np.ravel(np.where(time >= float(start)))[0]
    x2 = np.ravel(np.where(time <= float(stop)))[-1]
    
    if len(trace) == 0:
        print('Leak failed: empty window')
        
    elif type(trace) == np.ndarray:
        leak = np.mean(trace[x1:x2])
        no_leak = trace-leak
        
    elif type(trace) == list:
        leak = [np.mean(trace[i][x1:x2]) for i in range(len(trace))]
        no_leak = [trace[i]-leak[i] for i in range(len(trace))]
    return no_leak



def residual_sublimation(time, trace, start, stop, freq, n_peaks):
    ''' 
    Suppresses the residual preceding the onset of the peak from the peak
    time (array): time variable.
    trace (array or list): the trace or list of traces.
    start (int or float) : first limit of the window.
    end (int or float): second limit of the window.

    Returns
    -------
    no_res : list
        the list of traces containing peaks without res.

    '''
    
    if len(trace) == 0:
        print('Res sublimation failed: empty window')
        
    elif type(trace) == np.ndarray:
        x1 = np.ravel(np.where(time >= start))[0]
        x2 = np.ravel(np.where(time <= stop))[-1]
        x2_bis = np.ravel(np.where(time <= stop))[-1]
        bsl = trace[0:x2]
        PEAKS_NO_RES = []
        for i in range(n_peaks):
            res = np.mean(trace[x1:x2])
            if freq == '20':
                stop += 0.05
                x2_bis = np.ravel(np.where(time <= stop))[-1]
                peak_no_res = trace[x2:x2_bis]-res
                start += 0.05
            elif freq == '50':  
                stop += 0.02
                x2_bis = np.ravel(np.where(time <= stop))[-1]
                peak_no_res = trace[x2:x2_bis]-res
                start += 0.02
            elif freq == '100':  
                stop += 0.01
                x2_bis = np.ravel(np.where(time <= stop))[-1]
                peak_no_res = trace[x2:x2_bis]-res
                start += 0.01
                
            x1 = np.ravel(np.where(time >= start))[0]
            x2 = np.ravel(np.where(time <= stop))[-1]
            PEAKS_NO_RES.append(peak_no_res)
            
        end = np.append(np.hstack(PEAKS_NO_RES), trace[x2:])
        no_res = np.append(bsl, end)
        
    elif type(trace) == list:
        x2_bsl = np.ravel(np.where(time <= stop))[-1]
        no_res = []
        for i in range(len(trace)):
            alpha = start
            omega = stop
            x1 = np.ravel(np.where(time >= alpha))[0]
            x2 = np.ravel(np.where(time <= omega))[-1]
            x2_bis = np.ravel(np.where(time <= omega))[-1]
            bsl = trace[i][0:x2_bsl]
            PEAKS_NO_RES = []
            for j in range(n_peaks):
                res = np.mean(trace[i][x1:x2])
                if freq == '20':
                    omega += 0.05
                    x2_bis = np.ravel(np.where(time <= omega))[-1]
                    peak_no_res = trace[i][x2:x2_bis]-res
                    alpha += 0.05
                elif freq == '50':  
                    omega += 0.02
                    x2_bis = np.ravel(np.where(time <= omega))[-1]
                    peak_no_res = trace[i][x2:x2_bis]-res
                    alpha += 0.02
                elif freq == '100':
                    omega += 0.01
                    x2_bis = np.ravel(np.where(time <= omega))[-1]
                    peak_no_res = trace[i][x2:x2_bis]-res
                    alpha += 0.01
                
                x1 = np.ravel(np.where(time >= alpha))[0]
                x2 = np.ravel(np.where(time <= omega))[-1]
                PEAKS_NO_RES.append(peak_no_res)
 
            end = np.append(np.hstack(PEAKS_NO_RES), trace[i][x2:])
            no_res_trace = np.append(bsl, end)
            no_res.append(no_res_trace)
        
    return no_res
  
    
  
    
def values_extraction(time, trace, start, stop):    
    '''
    Exctracts values on a window of a dataset values
    time (array): time variable.
    trace (array or list): the trace or list of traces.
    start (int or float) : first limit of the window.
    end (int or float): second limit of the window.

    Returns
    -------
    windows : list
        The list of values in the extracted window.

    '''
    
    x1 = np.ravel(np.where(time >= float(start)))[0]
    x2 = np.ravel(np.where(time <= float(stop)))[-1]
    
    if len(trace) == 0:
        print('Extraction failed: empty window')
        
    elif type(trace) == np.ndarray:
        windows = trace[x1:x2]
        
    elif type(trace) == list:
        windows = [trace[i][x1:x2] for i in range(len(trace))]
        
    return windows




if __name__ == '__main__':
    
    '''
    >> loads STP profiles of individual parallel fibers boutons STP profiles from *.xlsx files (see load_xlsx def)
        /!\ Use ImagingPY.py to get the excel file with STP profiles
    >> sorts bootstrap histograms observations between peaks and noise to retrieve failures
    >> creates two excel files with the original file name:
        - *_data.xlsx: sorts variables related to the peakand noise amplitudes, the percentage of failures.
        - *_histograms: sorts all the bootstrap values of noise and peak for each episode
    
    LIST OF OUTPUT FIGURES
        SWEEPS Plot
        NOISE/PEAK Histograms
        % FAILURES/PEAK
        
    RETURNS
        - excel file (*_data.xlsx) with data related to amp, failures, z-score for each peak (1peak/sheet)
        - excel file (*_histogram.xlsx) with histograms values of noise and peak for each episode (1peak/sheet)
    
    '''
    
    import matplotlib.pyplot as plt
    from scipy.optimize import curve_fit
    from scipy.signal import savgol_filter
    import PySimpleGUI as sg
    import numpy as np
    import pandas as pd
    from scipy import stats
    import math
    import random
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    
    sg.theme('DarkBlue')

    layout = [[sg.Text('Excel File Name')],
              [sg.InputText(size=(35,1), key='Input'), sg.FileBrowse()],
              [sg.Checkbox('Show noise & peak#', default=False, key='Noise_peak'), sg.InputText(size=(3,1),default_text='1', key='Peak#')],
              [sg.Button('Start File')],
              [sg.Text('Saving folder')],
              [sg.InputText(size=(35,1), key='Save'), sg.FolderBrowse()],
              [sg.Frame(layout=[
              [sg.Text('Photobleaching'), sg.InputText(size=(4,1),default_text='0.01', key='Photo_start'), sg.Text('to'), sg.InputText(size=(4,1),default_text='0.45', key='Photo_stop'), sg.Text('sec'), sg.Button('Correction')],
              [sg.Text('Leak'), sg.InputText(size=(4,1),default_text='0.35', key='Leak_start'), sg.Text('to'), sg.InputText(size=(4,1),default_text='0.45', key='Leak_stop'), sg.Text('sec'), sg.Button('Subtraction')],
              [sg.Text('Residual'), sg.InputText(size=(5,1),default_text='0.49', key='Res_start'), sg.Text('to'), sg.InputText(size=(5,1),default_text='0.5', key='Res_stop'), sg.Text('sec'), sg.Button('Sublimation')]], title='Correction settings', relief=sg.RELIEF_SUNKEN)],
              [sg.Frame(layout=[
              [sg.Text('Freq (Hz)'), sg.InputText(size=(4,1),default_text='20', key='Frequency')],
              [sg.Text('Num peaks'), sg.InputText(size=(3,1),default_text='10', key='Peaks')],
              [sg.Text('Peak window'), sg.Text('Noise window')],
              [sg.InputText(size=(5,1),default_text='0.498', key='Peak_start'), sg.InputText(size=(5,1),default_text='0.51', key='Peak_stop'), sg.InputText(size=(5,1),default_text='0.1', key='Noise_start'), sg.InputText(size=(5,1),default_text='0.4', key='Noise_stop'), sg.Text('sec')],
              [sg.Checkbox('Show histograms', default=False, key='Fig_histograms')]], title='Windows', relief=sg.RELIEF_SUNKEN), sg.Button('GO')]]
             
                        
    window = sg.Window('BOOTSTRAP', layout, location=(0,0))
    
    while True:
        
        event, value = window.read()
        
        try:
            
            if event in (None, 'Close'):
                plt.close('all')
                break
                
            
            if event == 'Start File':
                correction = False
                subtraction = False
                sublimation = False
                
                time, avg, sweeps = load_xls(value['Input'])
                
                if value['Noise_peak'] == True:
                    fig = plt.figure(figsize=(10,3), tight_layout=True)
                    gs = fig.add_gridspec(1,3)
                    ax1 = fig.add_subplot(gs[:,0])
                else:
                    fig = plt.figure(figsize=(5,3), tight_layout=True)
                    gs = fig.add_gridspec(1,1)
                    ax1 = fig.add_subplot(gs[:,0])
                    
                if value['Noise_peak'] == True:
                    ax2 = fig.add_subplot(gs[:,1], sharey=ax1)
                    ax3 = fig.add_subplot(gs[:,2], sharey=ax1)
                
                title = str(value['Input']).split('/')
                ax1.set_title('{}'.format(title[-1]), fontsize=8)
                ax1.set_xlabel('Time(sec)')
                ax1.set_ylabel('dF/F0')
                if value['Noise_peak'] == True:
                    ax2.set_title('Noise')
                    ax3.set_title(f"Peak{value['Peak#']}")
                
                [ax1.plot(time, sweeps[i], 'k', alpha=0.1) for i in range(len(sweeps))]
                ax1.plot(time, avg, 'b', linewidth=2)
                ax1.plot([time[0],time[-1]], [0,0], 'k', linestyle='--')
                
                
            if event == 'Correction':
                correction = True
                ax1.clear()
                
                avg_unbleached = bleaching_correction(time, avg, value['Photo_start'], value['Photo_stop'])
                sweeps_unbleached = bleaching_correction(time, sweeps, value['Photo_start'], value['Photo_stop'])
                
                [ax1.plot(time, sweeps_unbleached[i], 'k', alpha=0.1) for i in range(len(sweeps_unbleached))]
                ax1.plot(time, avg_unbleached, 'b', linewidth=2, alpha=0.5)
                ax1.plot([time[0],time[-1]], [0,0], 'k', linestyle='--')
                
                title = str(value['Input']).split('/')
                ax1.set_title('{}'.format(title[-1]), fontsize=8)
                ax1.set_xlabel('Time(sec)')
                ax1.set_ylabel('dF/F0')
                if value['Noise_peak'] == True:
                    ax2.set_title('Noise')
                    ax3.set_title(f"Peak{int(value['Peak#'])}")
                
            
            if event == 'Substraction':
                subtraction = True
                ax1.clear()
                
                if correction == True:
                    avg_unleaked = leak_subtraction(time, avg_unbleached, value['Leak_start'], value['Leak_stop'])
                    sweeps_unleaked = leak_subtraction(time, sweeps_unbleached, value['Leak_start'], value['Leak_stop'])
                
                else:
                    avg_unleaked = leak_subtraction(time, avg, value['Leak_start'], value['Leak_stop'])
                    sweeps_unleaked = leak_subtraction(time, sweeps, value['Leak_start'], value['Leak_stop'])
                
                [ax1.plot(time, sweeps_unleaked[i], 'k', alpha=0.1) for i in range(len(sweeps_unleaked))]
                ax1.plot(time, avg_unleaked, 'b', linewidth=2, alpha=0.5)
                ax1.axhline(y=0.0, color='k', linestyle='--')
                
                title = str(value['Input']).split('/')
                ax1.set_title('{}'.format(title[-1]), fontsize=8)
                ax1.set_xlabel('Time(sec)')
                ax1.set_ylabel('dF/F0')
                if value['Noise_peak'] == True:
                    ax2.set_title('Noise')
                    ax3.set_title(f"Peak{int(value['Peak#'])}")
            
            
            if event == 'Sublimation':
                sublimation = True
                ax1.clear()
                
                if correction == True:
                    avg_sublimated = residual_sublimation(time, avg_unbleached, float(value['Res_start']), float(value['Res_stop']), str(value['Frequency']), int(value['Peaks']))
                    sweeps_sublimated = residual_sublimation(time, sweeps_unbleached, float(value['Res_start']), float(value['Res_stop']), str(value['Frequency']), int(value['Peaks']))
                
                if subtraction == True:
                    avg_sublimated = residual_sublimation(time, avg_unleaked, float(value['Res_start']), float(value['Res_stop']), str(value['Frequency']), int(value['Peaks']))
                    sweeps_sublimated = residual_sublimation(time, sweeps_unleaked, float(value['Res_start']), float(value['Res_stop']), str(value['Frequency']), int(value['Peaks']))
                
                else:
                    avg_sublimated = residual_sublimation(time, avg, float(value['Res_start']), float(value['Res_stop']), str(value['Frequency']), int(value['Peaks']))
                    sweeps_sublimated = residual_sublimation(time, sweeps, float(value['Res_start']), float(value['Res_stop']), str(value['Frequency']), int(value['Peaks']))
                
                [ax1.plot(time, sweeps_sublimated[i], 'k', alpha=0.1) for i in range(len(sweeps_sublimated))]
                ax1.plot(time, avg_sublimated, 'b', linewidth=2, alpha=0.5)
                ax1.axhline(y=0.0, color='k', linestyle='--')
                
                title = str(value['Input']).split('/')
                ax1.set_title('{}'.format(title[-1]), fontsize=8)
                ax1.set_xlabel('Time(sec)')
                ax1.set_ylabel('dF/F0')
                if value['Noise_peak'] == True:
                    ax2.set_title('Noise')
                    ax3.set_title(f"Peak{int(value['Peak#'])}")
                
                
            if event == 'GO':
                file_name = str(value['Input']).split('/')[-1].rsplit('.',1)[0]
                a = float(value['Peak_start'])
                b = float(value['Peak_stop'])
                
                WINDOWS_NS, WINDOWS_AMP = [],[]
                for i in range(int(value['Peaks'])):
                    if correction == True:
                        windows_amp = values_extraction(time, sweeps_unbleached, a, b)
                        windows_ns = values_extraction(time, sweeps_unbleached, value['Noise_start'], value['Noise_stop'])
                        
                    elif subtraction == True:
                        windows_amp = values_extraction(time, sweeps_unleaked, a, b)
                        windows_ns = values_extraction(time, sweeps_unleaked, value['Noise_start'], value['Noise_stop'])
                    
                    elif sublimation == True:
                        windows_amp = values_extraction(time, sweeps_sublimated, a, b)
                        windows_ns = values_extraction(time, sweeps_sublimated, value['Noise_start'], value['Noise_stop'])
                        
                    else:
                        windows_amp = values_extraction(time, sweeps, a, b)
                        windows_ns = values_extraction(time, sweeps, value['Noise_start'], value['Noise_stop'])
                    
                    ax1.axvline(x=a, color='cyan', linestyle='--')
                    ax1.axvline(x=b, color='cyan', linestyle='--')
                    
                    if str(value['Frequency']) == '20':
                        a += 0.05
                        b += 0.05
                    elif str(value['Frequency']) == '50':  
                        a += 0.02
                        b += 0.02
                    elif str(value['Frequency']) == '100':  
                        a += 0.01
                        b += 0.01
                        
                    WINDOWS_NS.append(windows_ns)
                    WINDOWS_AMP.append(windows_amp)
                    
              
                ax1.axvline(x=float(value['Noise_start']), color='k', linestyle='--')
                ax1.axvline(x=float(value['Noise_stop']), color='k', linestyle='--')
                
                if value['Noise_peak'] == True:
                    [ax2.plot(windows_ns[i], 'k', alpha=0.1) for i in range(len(windows_ns))]
                    [ax3.plot(WINDOWS_AMP[int(value['Peak#'])-1][i], 'k', alpha=0.1) for i in range(len(WINDOWS_AMP[int(value['Peak#'])-1]))]
                    ax2.axhline(y=0, color='k', linestyle='--')
                    ax3.axhline(y=0, color='k', linestyle='--')
                
                wb_hist = Workbook()
                wb_data = Workbook()

                fig3, ax3 = plt.subplots(figsize=(2,2), tight_layout=True)
                ALL_FAILS = []
                
                for peak in range(int(value['Peaks'])):
                    EP, PEAK, ZSCORE, AMP, NS_AMP, STD_NS, STD_AMP = [],[],[],[],[],[],[] 
                    
                    if value['Fig_histograms'] == True:
                        fig2, ax2 = plt.subplots(3, math.ceil(len(windows_ns)/3), figsize=(9,5), tight_layout=True)
                        fig2.suptitle('Bootstrap peak{}'.format(peak+1))
                    
                    df_episodes = pd.DataFrame(index=None, columns=None)
                    
                    i = 0
                    j = 0
                    
                    print('BOOTSTRAP PEAK{}'.format(peak+1))
                    print('----------------------')
                    print('----------------------')
                    
                    for item in range(len(WINDOWS_AMP[peak])):
                        EP.append('Episode {}'.format(item))
                    
                        amp_mean, amp_std, amp_cycle = bootstrap_patterns(WINDOWS_AMP[peak][item], N=len(WINDOWS_AMP[peak][item]))
                        
                        ns_cycle = [np.mean(random.sample(windows_ns[item].tolist(),len(WINDOWS_AMP[peak][item]))) for i in range(len(amp_cycle))]
                        ns_mean = np.mean(ns_cycle)
                        ns_std = np.std(ns_cycle)
                        # ns_mean, ns_std, ns_cycle = bootstrap_patterns(WINDOWS_NS[peak][item], N=len(WINDOWS_AMP[peak][item]))
                        
                        
                        NS_AMP.append(ns_mean)
                        STD_AMP.append(amp_std)
                        STD_NS.append(ns_std)
                        
                        amp = amp_mean - ns_mean
                        AMP.append(amp)
                        
                        df_episodes['noise{}'.format(item)] = ns_cycle
                        df_episodes['amp{}'.format(item)] = amp_cycle
                    
                        ns_n, ns_norm_pvalue = stats.shapiro(ns_cycle)
                        amp_n, amp_norm_pvalue = stats.shapiro(amp_cycle)
                        ns_l, ns_norm_l_pvalue = stats.levene(ns_cycle, amp_cycle)
                        
                        
                        print('Episode {}'.format(item))
                        print('Raw_mean: {:.3f} ; Bootstrap_mean: {:.3f}'.format(np.mean(windows_ns[item]), ns_mean))
                        print('Ns_mean: {:.3f} ; Ns_std: {:.3f}'.format(ns_mean, ns_std))
                        print('Amp_mean: {:.3f} ; Amp_std: {:.3f}'.format(amp_mean, amp_std))
                        print('Shapiro-Wilk test:\n noise_p = {:.3f} ; amp_p = {:.3f}'.format(ns_norm_pvalue, amp_norm_pvalue))
                        print('Levene test:\n p = {:.3f}'.format(ns_norm_l_pvalue))
                        
                        
                        if j == math.ceil((len(WINDOWS_AMP[peak])/3)):
                            i += 1
                            j = 0
                            
                        elif j == math.ceil((2*(len(WINDOWS_AMP[peak])/3))):
                            i += 1
                            j = 0
        
    
                        z_score = amp/ns_std
                        ZSCORE.append(z_score)
                        print('Z_SCORE:\n{:.3f}'.format(z_score))
                        
                        if z_score <= 3:
                            PEAK.append('Fail')
                            print('Not statistically significant')
                            if value['Fig_histograms'] == True:
                                ax2[i,j].text(0,0, 'zscore: {:.2f}'.format(z_score), fontsize=8)
                        else:
                            PEAK.append('Success')
                            print('Statistically significant')
                            if value['Fig_histograms'] == True:
                                ax2[i,j].text(0,0, 'zscore: {:.2f}'.format(z_score), fontsize=8, color='red')
                        
                        print('-------')
                        
                        
                        if value['Fig_histograms'] == True:
                            n_ns, bins_ns, patches_ns = ax2[i,j].hist(ns_cycle, bins=int(np.sqrt(df_episodes.shape[0])/2), density=True, facecolor='k', alpha=0.2)
                            n_amp, bins_amp, patches_amp = ax2[i,j].hist(amp_cycle, bins=int(np.sqrt(df_episodes.shape[0])/2), density=True, facecolor='cyan', alpha=0.4)
                            y_ns = gauss_func(ns_mean,ns_std,bins_ns)
                            y_amp = gauss_func(amp_mean,amp_std,bins_amp)
                            ax2[i,j].plot(bins_ns,y_ns,color='k',alpha=0.5)
                            ax2[i,j].plot(bins_amp,y_amp,color='cyan',alpha=0.6)
                            ax2[i,j].set_title('Episode {}\n NOISE µ:{:.2f}, σ:{:.2f}\n PEAK µ:{:.2f}, σ:{:.2f}'.format(item, ns_mean, ns_std, amp_mean, amp_std), fontsize=8)       
                        
                        j += 1
                
                    percentage_failures = (PEAK.count('Fail')/len(WINDOWS_AMP[peak]))*100
                    ALL_FAILS.append(percentage_failures)
                    
                    df_ep = pd.concat((pd.DataFrame(EP), pd.DataFrame(PEAK), pd.DataFrame(AMP), pd.DataFrame(NS_AMP),
                                       pd.DataFrame(STD_AMP), pd.DataFrame(STD_NS), pd.DataFrame(ZSCORE)),axis=1)
                    df_ep.columns = ['Episode','PEAK','AMP','AMPns','Std_amp','Std_ns','Zscore']
                    df_ep['PercFail'] = percentage_failures
                    
                    
                    sheet_hist = wb_hist.create_sheet('PEAK{}'.format(peak+1))
                    sheet_data = wb_data.create_sheet('PEAK{}'.format(peak+1))
                    
                    for r in dataframe_to_rows(df_episodes, index=False, header=True):
                        sheet_hist.append(r)
        
                    for r in dataframe_to_rows(df_ep, index=False, header=True):
                        sheet_data.append(r)

                    print('----------------------')
                    print('----------------------')
                
                
                ax3.plot(ALL_FAILS, color='r', marker='o', markerfacecolor='k')
                ax3.set_title('FAILURES')
                ax3.set_ylim(ymin=0, ymax=100)
                ax3.set_ylabel('% failure')
                ax3.set_xlabel('# Peak')
                
                
                wb_hist.remove(wb_hist['Sheet'])
                wb_data.remove(wb_data['Sheet'])
                
                wb_hist.save('{}\{}_histograms_bootstrap.xlsx'.format(value['Save'], file_name))
                wb_data.save('{}\{}_data_bootstrap.xlsx'.format(value['Save'], file_name))
                    
                sg.popup_ok('STORED')
                
        except:
            pass
        
    window.close()