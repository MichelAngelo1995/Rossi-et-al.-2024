# -*- coding: utf-8 -*-
"""
Created on Sat Nov 20 18:47:34 2021

@author: Theo.ROSSI
"""


def visualization(name, dataframe, save_path):
    
    time = dataframe['Time']
    episodes = dataframe.drop('Time', axis=1)
    average = np.mean(episodes, axis=1)
    
    f0_start = np.ravel(np.where(time >= 0.38))[0]
    f0_stop = np.ravel(np.where(time <= 0.48))[-1]
    
    f0_episodes = [np.mean(episodes.iloc[:,i][f0_start : f0_stop]) for i in range(len(episodes.columns))]
    f0_average = np.mean(average[f0_start : f0_stop])
    
    dF_F_traces = (episodes - f0_episodes)/f0_episodes
    dF_F_average = (average - f0_average)/f0_average
    
    fig, ax = plt.subplots(1,3, figsize=(12,4), tight_layout=True)
    
    for i in range(episodes.shape[1]):
        ax[0].plot(time, episodes.iloc[:,i], 'k', alpha=0.3)
        ax[1].plot(time, dF_F_traces.iloc[:,i], 'k', alpha=0.3)
        
    ax[0].plot(time, average)
    ax[0].set_xlabel('Time (sec)')
    ax[0].set_ylabel('A.U.')

    ax[1].set_ylabel('DF/F')
    
    ax[2].plot(time, savgol_filter(dF_F_average, 9, 2))
    
    df_traces = pd.concat((episodes, pd.DataFrame(average, columns=['Average']), time), axis=1)
    df_DF_F = pd.concat((dF_F_traces, pd.DataFrame(dF_F_average, columns=['Average']), time), axis=1)
    df_f0 = pd.DataFrame(f0_episodes).transpose()
    df_f0['Average'] = f0_average
    
    wb = Workbook()
    
    sheet1 = wb.create_sheet('Traces no Fback')
    sheet2 = wb.create_sheet('Traces DF_F0')
    sheet3 = wb.create_sheet('F0')
        
    [sheet1.append(i) for i in dataframe_to_rows(df_traces, index=False, header=True)]
    [sheet2.append(i) for i in dataframe_to_rows(df_DF_F, index=False, header=True)]
    [sheet3.append(i) for i in dataframe_to_rows(df_f0, index=False, header=True)]

    print('----------------------')

    wb.remove(wb['Sheet'])
    wb.save(f'{save_path}/{name}_converted.xlsx')
    print('File saved')
    




def traces_selection(name, tab, save_path):
    
    '''
    
    tab: dict
        Dictionary with excel sheet names (str) as keys and DataFrame of variables as values.
    
    Returns:    df_variables: DataFrame of shape (n_samples, n_variables)
    
    '''
    
    Checkboxes = {}
    
    Checkboxes['Traces'] = []
    
    for i in range(len(tab['Traces'].columns)):
        if tab['Traces'].columns[i] == 'Time':
            pass
        else:
            Checkboxes['Traces'].append([sg.Checkbox('Ep {}'.format(tab['Traces'].columns[i]), key='{}'.format(tab['Traces'].columns[i]))])
    
    sg.theme('DarkBlue')
    
    tab_layout = [[sg.InputText(default_text='2', size=(2,1), key='threshold'), sg.Text('Threshold'), sg.Button('Check_traces')],
                  [sg.Frame(layout=[
                  [sg.Checkbox('ALL', key='all_traces')],
                  [sg.TabGroup([[sg.Tab(sheet, checkbox_list) for sheet, checkbox_list in Checkboxes.items()]])]], title='EPISODES SELECTION', relief=sg.RELIEF_SUNKEN)],
                  [sg.Button('ANALYSE')]]
    
    window = sg.Window('VARIABLES').Layout([[sg.Column(tab_layout, size=(300,900), scrollable=True)]])
    
    while True:
        
        event, value = window.read()
        
        try:
            if event in (None, 'Close'):
                break
                
            
            if event == 'Check_traces':
                
                df_variables = pd.DataFrame(index=None)
                
                time = tab['Traces']['Time']
                traces_only = tab['Traces'].drop('Time', axis=1)
                
                for col in range(len(traces_only.columns)):
                    df_variables[f'{traces_only.columns[col]}'] = traces_only.iloc[:,col] - tab['Fback'][col]
                
                f0_start = np.ravel(np.where(time >= 0.38))[0]
                f0_stop = np.ravel(np.where(time <= 0.48))[-1]
                
                f0_ep = [np.mean(df_variables.iloc[:,i][f0_start : f0_stop]) for i in range(len(df_variables.columns))]
                f0_mean = np.mean(f0_ep)
                f0_std = np.std(f0_ep)
            
                fig, ax = plt.subplots(1,3,figsize=(14,4),tight_layout=True)
                [ax[0].plot(time, df_variables.iloc[:,i], alpha = 0.3, label=f'Ep{i}') for i in range(len(traces_only.columns))]
                ax[0].legend()
                sns.violinplot(data=f0_ep, ax=ax[1])
                
                threshold = int(value['threshold'])
                for i in range(len(f0_ep)):
                    z_score = (f0_ep[i] - f0_mean)/f0_std
                    print(z_score)
                    if z_score > threshold or z_score < -threshold:
                        ax[1].scatter(0, f0_ep[i], color = 'r')
                        ax[1].text(0, f0_ep[i], f'Ep{i}', color = 'r', fontsize = 12)
                        ax[2].scatter(i+1, z_score, color='r')
                    else:
                        ax[1].scatter(0, f0_ep[i], color = 'k')
                        ax[1].text(0, f0_ep[i], f'Ep{i}', color = 'k', fontsize = 12)
                        ax[2].scatter(i+1, z_score, color='k')
                
                print('-----------')
                
                ax[1].axhline(y=f0_mean, color='k', ls='--')
                ax[1].axhline(y=f0_mean+(threshold*f0_std), color='r', ls='--')
                ax[1].axhline(y=f0_mean-(threshold*f0_std), color='r', ls='--')
                
                ax[2].axhline(y=0, color='k', ls='--')
                ax[2].axhline(y=threshold, color='r', ls='--')
                ax[2].axhline(y=-threshold, color='r', ls='--')
                
                
                
            if event == 'ANALYSE':
                
                df_variables = pd.DataFrame(index=None)
                
                if value['all_traces'] == True:
                    
                    ep = tab['Traces'].drop('Time', axis=1)
                    
                    for col in range(len(ep.columns)):
                        
                        df_variables[f'{ep.columns[col]}'] = ep.iloc[:,col] - tab['Fback'][col]
                
                for sheet, checkbox_list in Checkboxes.items():
                    
                    for item in range(len(checkbox_list)):
                        
                        if value[f'{tab[sheet].columns[item]}'] == True:
                            
                            df_variables[f'{tab[sheet].columns[item]}'] = tab['Traces'].iloc[:,item] - tab['Fback'][item]                        
                            
            
                df_variables['Time'] = tab['Traces']['Time']
                
                print('Traces without Fback')
                print('------------------')
                print(df_variables)
                visualization(name, df_variables, save_path)
                
        except:
            pass
    window.close()
    
    return(df_variables)




if __name__ == '__main__':
    
    import PySimpleGUI as sg
    import pandas as pd
    import numpy as np
    import seaborn as sns
    import matplotlib.pyplot as plt
    from scipy.signal import savgol_filter
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows


    sg.theme('DarkBlue')
        
    main_layout = [[sg.Text('File path')],
                   [sg.InputText(size=(35,1), key='path'), sg.FileBrowse()],
                   [sg.Text('Saving path')],
                   [sg.InputText(size=(35,1), key='save'), sg.FolderBrowse()],
                   [sg.Button('GO')]]
              
    main_window = sg.Window('DF_F_conversion', main_layout, location=(0,0))
                   
    
    while True:
       main_event, main_value = main_window.read()
       try:
           if main_event in (None, 'Close'):
               plt.close('all')
               break
           
           if main_event == 'GO':
               name = main_value['path'].rsplit('/', 1)[-1].rsplit('.', 1)[0]
               Dataset = {}
               
               traces = pd.read_excel(f"{main_value['path']}", sheet_name='Traces', header=0)
               f_back = pd.read_excel(f"{main_value['path']}", sheet_name='Data', header=0).iloc[-1,1:]
               
               Dataset['Traces'] = traces
               Dataset['Fback'] = f_back
               
               traces_selection(name, Dataset, main_value['save'])
               
              
       except:
           pass
    main_window.close()